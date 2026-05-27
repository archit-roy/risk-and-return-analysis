import pandas as pd
import numpy as np
import sqlite3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from analysis import load_data, compute_returns, compute_metrics, compute_correlation, compute_portfolio

DARK      = "0B0B0D"
ACCENT    = "2563EB"
GREEN     = "16A34A"
RED       = "DC2626"
YELLOW    = "D97706"
WHITE     = "F8FAFC"
MID       = "94A3B8"
HEADER_BG = "1E293B"

def border():
    s = Side(border_style="thin", color="CBD5E1")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(color):    return PatternFill("solid", fgColor=color)
def bw(sz=11):      return Font(bold=True, color=WHITE, size=sz)
def center():       return Alignment(horizontal="center", vertical="center")
def cw(ws,col,w):   ws.column_dimensions[get_column_letter(col)].width = w

def header_row(ws, row, cols, texts):
    for col, txt in zip(cols, texts):
        c = ws.cell(row=row, column=col, value=txt)
        c.font = bw(); c.fill = fill(HEADER_BG)
        c.alignment = center(); c.border = border()

def sheet1(wb, metrics, port):
    ws = wb.create_sheet("Risk Metrics")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:J1")
    ws["A1"].value = "Portfolio Risk & Return — Indian Equities (2022–2025)"
    ws["A1"].font = Font(bold=True, color=WHITE, size=14)
    ws["A1"].fill = fill(DARK); ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 32
    ws.merge_cells("A2:J2")
    ws["A2"].value = "Equal-Weight 10-Stock Portfolio  |  Risk-Free Rate: 6.5% (RBI Repo)"
    ws["A2"].font = Font(color=MID, size=10, italic=True)
    ws["A2"].fill = fill("0F172A"); ws["A2"].alignment = center()

    cols = list(range(1,11))
    headers = ["Ticker","Start ₹","End ₹","Total Ret %","Ann Ret %","Ann Vol %","Sharpe","Max DD %","Win Rate %","Days"]
    header_row(ws, 3, cols, headers)
    for col, w in zip(cols, [12,10,10,12,12,11,10,11,11,8]):
        cw(ws, col, w)

    for ri, row in enumerate(metrics.itertuples(), start=4):
        bg = fill("F1F5F9") if ri%2==0 else fill(WHITE)
        vals = [row.ticker, row.start_price, row.end_price, row.total_return_pct,
                row.ann_return_pct, row.ann_volatility_pct, row.sharpe_ratio,
                row.max_drawdown_pct, row.win_rate_pct, row.trading_days]
        for col, val in zip(cols, vals):
            c = ws.cell(row=ri, column=col, value=val)
            c.fill = bg; c.alignment = center(); c.border = border()
            c.font = Font(size=10, color="1E293B")
            if col == 5:
                c.font = Font(size=10, bold=True, color=GREEN if val>0 else RED)
            if col == 8:
                c.font = Font(size=10, bold=True, color=RED)
            if col == 7:
                color = GREEN if val>=1 else (YELLOW if val>=0 else RED)
                c.font = Font(size=10, bold=True, color=color)
        ws.row_dimensions[ri].height = 18

    pr = ri + 2
    ws.merge_cells(start_row=pr, start_column=1, end_row=pr, end_column=10)
    ws.cell(row=pr, column=1, value="EQUAL-WEIGHT PORTFOLIO SUMMARY")
    ws.cell(row=pr, column=1).font = bw(12)
    ws.cell(row=pr, column=1).fill = fill(ACCENT)
    ws.cell(row=pr, column=1).alignment = center()
    ph = ["Portfolio","Total Ret %","Ann Ret %","Ann Vol %","Sharpe","Max DD %"]
    pv = [port["portfolio"], port["total_return_pct"], port["ann_return_pct"],
          port["ann_volatility_pct"], port["sharpe_ratio"], port["max_drawdown_pct"]]
    for i,(h,v) in enumerate(zip(ph,pv)):
        hc = ws.cell(row=pr+1, column=i+1, value=h)
        hc.font=bw(); hc.fill=fill(HEADER_BG); hc.alignment=center(); hc.border=border()
        vc = ws.cell(row=pr+2, column=i+1, value=v)
        vc.font=Font(size=11,bold=True,color=DARK); vc.fill=fill("F1F5F9")
        vc.alignment=center(); vc.border=border()

def sheet2(wb, metrics):
    ws = wb.create_sheet("Sharpe Ranking")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"].value = "Sharpe Ratio Ranking — Risk-Adjusted Returns"
    ws["A1"].font = Font(bold=True, color=WHITE, size=13)
    ws["A1"].fill = fill(DARK); ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30
    header_row(ws, 2, range(1,7), ["Rank","Ticker","Ann Ret %","Ann Vol %","Sharpe","Signal"])

    ranked = metrics.sort_values("sharpe_ratio", ascending=False).reset_index(drop=True)
    for ri, row in enumerate(ranked.itertuples(), start=3):
        s = row.sharpe_ratio
        sig = "Strong Buy" if s>=1 else ("Buy" if s>=0.5 else ("Hold" if s>=0 else "Avoid"))
        sf = fill(GREEN) if "Buy" in sig else (fill(YELLOW) if sig=="Hold" else fill(RED))
        for col,(val,fl) in enumerate(zip([ri-2,row.ticker,row.ann_return_pct,
                                           row.ann_volatility_pct,s,sig],
                                          [None]*5+[sf]),1):
            c = ws.cell(row=ri, column=col, value=val)
            c.alignment=center(); c.border=border()
            if fl: c.fill=fl; c.font=bw(10)
            else:
                c.fill=fill("F1F5F9") if ri%2==0 else fill(WHITE)
                c.font=Font(size=10)
        ws.row_dimensions[ri].height = 18

    last = 2+len(ranked)
    chart = BarChart()
    chart.type="col"; chart.title="Sharpe Ratio by Stock"
    chart.y_axis.title="Sharpe Ratio"; chart.style=10
    chart.width=20; chart.height=12
    chart.add_data(Reference(ws,min_col=5,min_row=2,max_row=last),titles_from_data=True)
    chart.set_categories(Reference(ws,min_col=2,min_row=3,max_row=last))
    ws.add_chart(chart,f"A{last+3}")
    for col,w in zip(range(1,7),[6,12,12,12,10,12]): cw(ws,col,w)

def sheet3(wb, metrics):
    ws = wb.create_sheet("Drawdown Analysis")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:E1")
    ws["A1"].value = "Maximum Drawdown Analysis — Capital Preservation View"
    ws["A1"].font = Font(bold=True, color=WHITE, size=13)
    ws["A1"].fill = fill(DARK); ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30
    header_row(ws,2,range(1,6),["Ticker","Max Drawdown %","Ann Vol %","Win Rate %","Risk Level"])

    for ri,row in enumerate(metrics.sort_values("max_drawdown_pct").itertuples(),start=3):
        dd = row.max_drawdown_pct
        risk = "High Risk" if dd<-30 else ("Medium Risk" if dd<-20 else "Lower Risk")
        rf = fill(RED) if "High" in risk else (fill(YELLOW) if "Medium" in risk else fill(GREEN))
        for col,val in enumerate([row.ticker,dd,row.ann_volatility_pct,row.win_rate_pct,risk],1):
            c = ws.cell(row=ri, column=col, value=val)
            c.alignment=center(); c.border=border()
            if col==5: c.fill=rf; c.font=bw(10)
            elif col==2: c.fill=fill("F1F5F9"); c.font=Font(bold=True,color=RED,size=10)
            else:
                c.fill=fill("F1F5F9") if ri%2==0 else fill(WHITE)
                c.font=Font(size=10)
        ws.row_dimensions[ri].height = 18
    for col,w in zip(range(1,6),[12,16,12,12,14]): cw(ws,col,w)

def sheet4(wb, corr):
    ws = wb.create_sheet("Correlation Matrix")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:K1")
    ws["A1"].value = "Return Correlation Matrix — Diversification Analysis"
    ws["A1"].font = Font(bold=True, color=WHITE, size=13)
    ws["A1"].fill = fill(DARK); ws["A1"].alignment = center()
    ws.row_dimensions[1].height = 30
    tickers = list(corr.columns)
    ws.cell(row=2,column=1,value="").fill = fill(HEADER_BG)
    for i,t in enumerate(tickers,2):
        c=ws.cell(row=2,column=i,value=t)
        c.font=bw(); c.fill=fill(HEADER_BG); c.alignment=center(); c.border=border()
    for ri,tr in enumerate(tickers,start=3):
        rh=ws.cell(row=ri,column=1,value=tr)
        rh.font=bw(); rh.fill=fill(HEADER_BG); rh.alignment=center(); rh.border=border()
        for ci,tc in enumerate(tickers,start=2):
            val=corr.loc[tr,tc]
            c=ws.cell(row=ri,column=ci,value=round(val,3))
            c.alignment=center(); c.border=border()
            if tr==tc: c.fill=fill(ACCENT); c.font=bw()
            elif val>=0.8: c.fill=fill("FEE2E2"); c.font=Font(color=RED,bold=True,size=10)
            elif val>=0.5: c.fill=fill("FEF3C7"); c.font=Font(color="92400E",size=10)
            else: c.fill=fill("DCFCE7"); c.font=Font(color=GREEN,size=10)
        ws.row_dimensions[ri].height = 18
    for col in range(1,len(tickers)+2): cw(ws,col,13)

if __name__ == "__main__":
    df = load_data()
    df_ret = compute_returns(df)
    metrics = compute_metrics(df_ret)
    corr = compute_correlation(df_ret)
    port = compute_portfolio(df_ret)

    wb = Workbook()
    wb.remove(wb.active)
    sheet1(wb, metrics, port)
    sheet2(wb, metrics)
    sheet3(wb, metrics)
    sheet4(wb, corr)
    wb.save("portfolio_risk_return_dashboard.xlsx")
    print("Saved → portfolio_risk_return_dashboard.xlsx")